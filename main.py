import os
import io
import csv
import json
import logging
import asyncio
from datetime import datetime, timedelta
from typing import List, Optional
from contextlib import asynccontextmanager

import httpx
from fastapi import (
    FastAPI,
    Depends,
    HTTPException,
    Request,
    status,
    UploadFile,
    File,
    Form,
    Response,
)
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import PlainTextResponse, JSONResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy import select, func, text, update

from database import engine, Base, get_db, AsyncSessionLocal
from models import (
    User,
    Participant,
    Task,
    TasksToParticipant,
    CheckTasks,
    FormatFile,
    Notification,
    AboutContest,
    News,
    FAQ,
)
import auth
import vk_service
import schemas

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


async def deadline_notification_loop():
    """Каждый час проверяет задания с дедлайном через ~24 часа и уведомляет участников."""
    while True:
        try:
            async with AsyncSessionLocal() as db:
                now = datetime.utcnow()
                window_start = now + timedelta(hours=20)
                window_end = now + timedelta(hours=28)
                tasks_res = await db.execute(
                    select(Task).where(
                        Task.is_active == True,
                        Task.deadline_tasks >= window_start,
                        Task.deadline_tasks <= window_end,
                    )
                )
                tasks = tasks_res.scalars().all()

                for task in tasks:
                    notif_title = f"Дедлайн задания через 1 день"
                    notif_msg = f'До дедлайна задания «{task.name_tasks}» остался 1 день. Успейте сдать работу!'

                    participants_res = await db.execute(
                        select(User).where(User.role_user == "participant")
                    )
                    participants = participants_res.scalars().all()

                    for user in participants:
                        existing = await db.execute(
                            select(Notification).where(
                                Notification.id_user == user.id_user,
                                Notification.title == notif_title,
                                Notification.message == notif_msg,
                            )
                        )
                        if existing.scalar_one_or_none():
                            continue
                        db.add(Notification(
                            id_user=user.id_user,
                            title=notif_title,
                            message=notif_msg,
                        ))
                    await db.commit()
        except Exception as e:
            print(f"Deadline notification error: {e}")
        await asyncio.sleep(3600)


# Lifespan
@asynccontextmanager
async def lifespan(app: FastAPI):
    print("Проверка подключения к базе данных...")

    try:
        async with engine.connect() as conn:
            await conn.execute(text("SELECT 1"))
            print("База данных доступна")
    except Exception as e:
        print(f"Ошибка подключения: {e}")
        raise

    task = asyncio.create_task(deadline_notification_loop())

    yield

    task.cancel()
    await engine.dispose()
    print("Соединения закрыты")


app = FastAPI(
    title="Медиаконкурс",
    version="2.0.0",
    lifespan=lifespan,
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)


# ========== ОБРАБОТЧИКИ ОШИБОК ==========
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    logger.error(f"Ошибка валидации: {exc}")
    return JSONResponse(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        content={
            "detail": "Ошибка в данных запроса",
            "errors": exc.errors(),
        },
    )


# ========== АУТЕНТИФИКАЦИЯ ==========
@app.post("/api/register", tags=["Auth"])
async def register(user_data: schemas.UserCreate, db: AsyncSession = Depends(get_db)):
    """Регистрация нового пользователя"""
    # Проверка существующего пользователя
    result = await db.execute(select(User).where(User.login_user == user_data.username))
    if result.scalar_one_or_none():
        raise HTTPException(
            status_code=400, detail="Пользователь с таким логином уже существует."
        )

    result = await db.execute(select(User).where(User.email_user == user_data.email))
    if result.scalar_one_or_none():
        raise HTTPException(
            status_code=400, detail="Пользователь с таким email уже существует."
        )

    # Пробуем получить числовой VK ID из ссылки при регистрации
    vk_id_resolved = None
    if user_data.vk_link:
        try:
            async with httpx.AsyncClient(timeout=8.0) as client:
                vk_id_resolved = await vk_service.extract_user_id(user_data.vk_link, client)
        except Exception:
            pass

    # Проверка дублирующейся VK-ссылки (по числовому ID)
    if vk_id_resolved:
        result = await db.execute(select(User).where(User.vk_id == vk_id_resolved))
        if result.scalar_one_or_none():
            raise HTTPException(
                status_code=400,
                detail="Такой аккаунт уже существует. Войдите или восстановите пароль."
            )

    # Проверка дублирующейся VK-ссылки (точное совпадение, на случай недоступности API)
    result = await db.execute(
        select(Participant).where(Participant.vk_participant == user_data.vk_link)
    )
    if result.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail="Такой аккаунт уже существует. Войдите или восстановите пароль."
        )

    # Создаём пользователя
    new_user = User(
        name_user=user_data.full_name,
        login_user=user_data.username,
        email_user=user_data.email,
        password_user=auth.get_password_hash(user_data.password),
        role_user="participant",
        vk_id=vk_id_resolved,
    )
    db.add(new_user)
    await db.flush()

    # Создаём запись участника
    new_participant = Participant(
        id_user=new_user.id_user,
        vk_participant=user_data.vk_link,
        tg_participant=user_data.tg_link if user_data.tg_link else "",
        institute_participant=user_data.institute,
    )
    db.add(new_participant)

    await db.commit()
    await db.refresh(new_user)

    return {
        "id": new_user.id_user,
        "username": new_user.login_user,
        "full_name": new_user.name_user,
        "email": new_user.email_user,
        "role": new_user.role_user,
        "institute": new_participant.institute_participant,
        "vk_link": new_participant.vk_participant,
        "tg_link": new_participant.tg_participant,
    }


@app.post("/api/login", tags=["Auth"])
async def login(credentials: schemas.UserLogin, db: AsyncSession = Depends(get_db)):
    """Вход в систему"""
    result = await db.execute(
        select(User).where(User.login_user == credentials.username)
    )
    user = result.scalar_one_or_none()

    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден. Пожалуйста, зарегистрируйтесь.")
    if not auth.verify_password(credentials.password, user.password_user):
        raise HTTPException(status_code=401, detail="Неверный пароль.")

    access_token = auth.create_access_token(data={"sub": user.login_user})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.role_user,
        "full_name": user.name_user,
    }


async def _find_user_by_vk_id(vk_id_str: str, db: AsyncSession):
    """Находит пользователя платформы по числовому VK ID."""
    # Ищем по сохранённому числовому vk_id
    result = await db.execute(select(User).where(User.vk_id == vk_id_str))
    user = result.scalar_one_or_none()

    # Если не нашли — пробуем сопоставить по ссылке VK (для ранее зарегистрированных)
    if not user:
        parts_result = await db.execute(
            select(Participant).where(Participant.vk_participant.isnot(None))
        )
        participants = parts_result.scalars().all()
        async with httpx.AsyncClient(timeout=10.0) as client:
            for participant in participants:
                if not participant.vk_participant:
                    continue
                resolved = await vk_service.extract_user_id(participant.vk_participant, client)
                if resolved == vk_id_str:
                    user_result = await db.execute(
                        select(User).where(User.id_user == participant.id_user)
                    )
                    user = user_result.scalar_one_or_none()
                    if user:
                        # Сохраняем vk_id для быстрого поиска в будущем
                        user.vk_id = vk_id_str
                        await db.commit()
                        break

    return user


async def _find_user_by_vk_token(access_token: str, db: AsyncSession):
    """Находит пользователя по VK access_token (для сброса пароля)."""
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(
                "https://api.vk.com/method/users.get",
                params={"access_token": access_token, "v": "5.199"},
            )
            data = resp.json()
        if "error" in data:
            return None
        vk_id_str = str(data["response"][0]["id"])
    except Exception:
        return None

    return await _find_user_by_vk_id(vk_id_str, db)


@app.post("/api/auth/vk", tags=["Auth"])
async def auth_vk(
    vk_data: schemas.VKAuth,
    db: AsyncSession = Depends(get_db),
):
    """Вход через VK ID"""
    # user_id уже получен от VKID SDK при обмене кода — используем напрямую,
    # без лишнего запроса к VK API (users.get с новыми токенами VKID SDK ненадёжен)
    vk_id_str = str(vk_data.user_id)
    user = await _find_user_by_vk_id(vk_id_str, db)
    if not user:
        raise HTTPException(
            status_code=404,
            detail="Аккаунт не найден. Пожалуйста, зарегистрируйтесь на платформе."
        )
    access_token = auth.create_access_token(data={"sub": user.login_user})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.role_user,
        "full_name": user.name_user,
    }


@app.post("/api/reset-password", tags=["Auth"])
async def reset_password_vk(
    data: schemas.ResetPasswordVK,
    db: AsyncSession = Depends(get_db),
):
    """Сброс пароля через VK ID"""
    user = await _find_user_by_vk_token(data.access_token, db)
    if not user:
        raise HTTPException(
            status_code=404,
            detail="Аккаунт не найден. Убедитесь, что ВК аккаунт привязан к платформе."
        )
    user.password_user = auth.get_password_hash(data.new_password)
    await db.commit()
    return {"message": "Пароль успешно изменён"}


@app.get("/api/users/me", tags=["Auth"])
async def get_current_user_profile(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_user),
):
    """Получение данных текущего пользователя"""
    result = await db.execute(
        select(Participant).where(Participant.id_user == current_user.id_user)
    )
    participant = result.scalar_one_or_none()

    return {
        "id": current_user.id_user,
        "username": current_user.login_user,
        "full_name": current_user.name_user,
        "email": current_user.email_user,
        "role": current_user.role_user,
        "institute": participant.institute_participant if participant else None,
        "vk_link": participant.vk_participant if participant else None,
        "tg_link": participant.tg_participant if participant else None,
    }


@app.get("/api/users/me/history", tags=["User"])
async def get_user_history(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_user),
):
    """Получение истории выполненных заданий пользователя"""
    result = await db.execute(
        select(TasksToParticipant)
        .options(
            selectinload(TasksToParticipant.task),
            selectinload(TasksToParticipant.check),
        )
        .where(TasksToParticipant.id_user == current_user.id_user)
        .order_by(TasksToParticipant.date_task_to_participant.desc())
    )
    submissions = result.scalars().all()

    history = []
    for s in submissions:
        score = s.check.points_check_tasks if s.check else 0
        history.append(
            {
                "id": s.id_tasks_to_participant,
                "task_id": s.id_tasks,
                "task_title": s.task.name_tasks if s.task else f"Задание #{s.id_tasks}",
                "submission_data": s.url_participant,
                "status": s.status,
                "score": score,
                "submitted_at": (
                    s.date_task_to_participant.isoformat()
                    if s.date_task_to_participant
                    else None
                ),
            }
        )

    return history


# ========== УПРАВЛЕНИЕ ЗАДАНИЯМИ ==========
@app.post("/api/tasks", tags=["Admin/Tasks"])
async def create_task(
    task_data: schemas.TaskCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Создание задания (только для админов)"""
    try:
        # Рассчитываем баллы для auto заданий
        points_to_save = task_data.points_at_stake
        if task_data.task_type == "auto" and task_data.posts_urls:
            num_posts = len(task_data.posts_urls)
            multiplier = 1 if task_data.auto_type == "likes" else 3
            points_to_save = num_posts * multiplier

        new_task = Task(
            id_user=current_user.id_user,
            name_tasks=task_data.title,
            description_tasks=task_data.description,
            type_tasks=0 if task_data.task_type == "auto" else 1,
            deadline_tasks=task_data.deadline,
            is_active=True,
            auto_type=task_data.auto_type if task_data.task_type == "auto" else None,
            posts_urls=(
                json.dumps(task_data.posts_urls) if task_data.posts_urls else None
            ),
            points_per_action=points_to_save,
        )
        db.add(new_task)
        await db.flush()

        # Сохраняем форматы файлов для ручных заданий
        if task_data.task_type == "manual" and task_data.format_type:
            format_file = FormatFile(
                id_tasks=new_task.id_tasks, name_format_file=task_data.format_type
            )
            db.add(format_file)

        await db.commit()
        await db.refresh(new_task)

        # Уведомление всех участников
        result = await db.execute(select(User).where(User.role_user == "participant"))
        users = result.scalars().all()

        for user in users:
            notification = Notification(
                id_user=user.id_user,
                title="Новое задание!",
                message=f"Опубликовано задание «{new_task.name_tasks}». Максимальный балл: {points_to_save}",
            )
            db.add(notification)

        await db.commit()

        return {
            "message": "Задание успешно создано",
            "calculated_points": points_to_save,
        }

    except Exception as e:
        print(f"Ошибка создания задания: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/tasks", tags=["Admin/Tasks"])
async def get_all_tasks(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Получение всех заданий для админа"""
    result = await db.execute(select(Task).order_by(Task.id_tasks.desc()))
    tasks = result.scalars().all()

    return [
        {
            "id": t.id_tasks,
            "title": t.name_tasks,
            "description": t.description_tasks,
            "task_type": "auto" if t.type_tasks == 0 else "manual",
            "auto_type": t.auto_type,
            "points_at_stake": t.points_per_action,
            "deadline": t.deadline_tasks,
            "is_active": t.is_active,
            "posts_urls": json.loads(t.posts_urls) if t.posts_urls else [],
        }
        for t in tasks
    ]


@app.get("/api/tasks/available", tags=["User/Tasks"])
async def list_available_tasks(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_user),
):
    """Доступные задания для пользователя"""
    now = datetime.utcnow()

    # Получаем ID выполненных заданий
    result = await db.execute(
        select(TasksToParticipant.id_tasks).where(
            TasksToParticipant.id_user == current_user.id_user
        )
    )
    submitted_ids = [row[0] for row in result.all()]

    # Активные задания с будущим дедлайном
    query = select(Task).where(Task.is_active == True, Task.deadline_tasks > now)
    if submitted_ids:
        query = query.where(~Task.id_tasks.in_(submitted_ids))

    result = await db.execute(query)
    tasks = result.scalars().all()

    return [
        {
            "id": t.id_tasks,
            "title": t.name_tasks,
            "description": t.description_tasks,
            "task_type": "auto" if t.type_tasks == 0 else "manual",
            "auto_type": t.auto_type,
            "points_at_stake": t.points_per_action,
            "deadline": t.deadline_tasks,
            "is_active": t.is_active,
            "posts_urls": json.loads(t.posts_urls) if t.posts_urls else [],
        }
        for t in tasks
    ]


@app.post("/api/tasks/{task_id}/submit", tags=["User/Tasks"])
async def submit_task_solution(
    task_id: int,
    submission_url: Optional[str] = Form(None),
    files: List[UploadFile] = File(default=[]),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_user),
):
    """Прием решения задачи"""
    result = await db.execute(select(Task).where(Task.id_tasks == task_id))
    task = result.scalar_one_or_none()

    if not task:
        raise HTTPException(status_code=404, detail="Задание не найдено.")

    if task.deadline_tasks < datetime.utcnow():
        raise HTTPException(status_code=400, detail="Дедлайн задания истек.")

    # Проверка на повторную сдачу
    result = await db.execute(
        select(TasksToParticipant).where(
            TasksToParticipant.id_tasks == task_id,
            TasksToParticipant.id_user == current_user.id_user,
        )
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Вы уже сдали это задание.")

    submission = TasksToParticipant(
        id_tasks=task_id,
        id_user=current_user.id_user,
        url_participant="",
        status="pending",
    )

    # Автоматическая проверка VK
    final_score = 0
    if task.type_tasks == 0:  # auto
        posts_list = json.loads(task.posts_urls) if task.posts_urls else []

        # Получаем VK ID участника
        part_result = await db.execute(
            select(Participant).where(Participant.id_user == current_user.id_user)
        )
        participant = part_result.scalar_one_or_none()
        vk_id = participant.vk_participant if participant else None

        if not vk_id:
            raise HTTPException(
                status_code=400,
                detail="В профиле не указана ссылка VK для проверки. Пожалуйста, обновите профиль.",
            )

        score = await vk_service.check_vk_activity(
            vk_id, posts_list, task.auto_type or "likes"
        )
        final_score = int(score)
        submission.status = "approved" if score > 0 else "rejected"
        submission.url_participant = f"auto_check_{task.auto_type}"

        db.add(submission)
        await db.flush()

        # Создаём запись проверки
        check = CheckTasks(
            id_submission=submission.id_tasks_to_participant,
            id_user=current_user.id_user,
            id_tasks=task_id,
            points_check_tasks=final_score,
        )
        db.add(check)

    else:  # manual
        valid_files = [f for f in files if f and f.filename]
        if valid_files:
            saved_paths = []
            ts = int(datetime.now().timestamp())
            for i, file in enumerate(valid_files):
                filename = file.filename.lower()
                ext = os.path.splitext(filename)[1]
                unique_name = f"user_{current_user.id_user}_task_{task_id}_{ts}_{i}{ext}"
                file_path = os.path.join(UPLOAD_DIR, unique_name)
                contents = await file.read()
                with open(file_path, "wb") as f:
                    f.write(contents)
                saved_paths.append(f"/uploads/{unique_name}")

            submission.url_participant = (
                json.dumps(saved_paths) if len(saved_paths) > 1 else saved_paths[0]
            )

        elif submission_url:
            submission.url_participant = submission_url
        else:
            raise HTTPException(
                status_code=400, detail="Необходимо прикрепить файл или указать ссылку."
            )

        db.add(submission)

    await db.commit()

    return {
        "status": submission.status,
        "score": final_score,
        "message": (
            "Работа отправлена на проверку"
            if submission.status == "pending"
            else "Работа проверена автоматически"
        ),
    }


@app.put("/api/tasks/{task_id}", tags=["Admin/Tasks"])
async def edit_task(
    task_id: int,
    task_data: schemas.TaskCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Редактирование задания"""
    result = await db.execute(select(Task).where(Task.id_tasks == task_id))
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail="Задание не найдено.")

    points_to_save = task_data.points_at_stake
    if task_data.task_type == "auto" and task_data.posts_urls:
        multiplier = 1 if task_data.auto_type == "likes" else 3
        points_to_save = len(task_data.posts_urls) * multiplier

    task.name_tasks = task_data.title
    task.description_tasks = task_data.description
    task.type_tasks = 0 if task_data.task_type == "auto" else 1
    task.deadline_tasks = task_data.deadline
    task.auto_type = task_data.auto_type if task_data.task_type == "auto" else None
    task.posts_urls = json.dumps(task_data.posts_urls) if task_data.posts_urls else None
    task.points_per_action = points_to_save

    # Обновляем формат файла
    await db.execute(
        update(FormatFile).where(FormatFile.id_tasks == task_id).values(
            name_format_file=task_data.format_type or "Ссылка"
        )
    )
    if task_data.task_type == "manual" and task_data.format_type:
        # Создаём если не было
        fmt_result = await db.execute(
            select(FormatFile).where(FormatFile.id_tasks == task_id)
        )
        if not fmt_result.scalar_one_or_none():
            db.add(FormatFile(id_tasks=task_id, name_format_file=task_data.format_type))

    await db.commit()
    return {"message": "Задание обновлено", "calculated_points": points_to_save}


@app.post("/api/admin/tasks/{task_id}/archive", tags=["Admin/Tasks"])
async def archive_task(
    task_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Архивация задания"""
    await db.execute(
        update(Task).where(Task.id_tasks == task_id).values(is_active=False)
    )
    await db.commit()
    return {"status": "success", "message": "Задание архивировано"}


@app.delete("/api/admin/tasks/{task_id}", tags=["Admin/Tasks"])
async def delete_task(
    task_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Удаление архивного задания"""
    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    if task.is_active:
        raise HTTPException(status_code=400, detail="Нельзя удалить активное задание. Сначала переведите в архив.")
    await db.delete(task)
    await db.commit()
    return {"status": "success", "message": "Задание удалено"}


# ========== ПРОВЕРКА РАБОТ ==========
@app.get("/api/admin/submissions", tags=["Admin"])
async def get_all_submissions(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Все отправленные работы для админа"""
    result = await db.execute(
        select(TasksToParticipant)
        .options(selectinload(TasksToParticipant.check))
        .order_by(TasksToParticipant.id_tasks_to_participant.desc())
    )
    submissions = result.scalars().all()

    # Загружаем связанные данные
    result_tasks = await db.execute(select(Task))
    tasks = {t.id_tasks: t for t in result_tasks.scalars().all()}

    result_users = await db.execute(select(User))
    users = {u.id_user: u for u in result_users.scalars().all()}

    result_participants = await db.execute(select(Participant))
    participants = {p.id_user: p for p in result_participants.scalars().all()}

    result_list = []
    for s in submissions:
        task = tasks.get(s.id_tasks)
        user = users.get(s.id_user)
        participant = participants.get(s.id_user) if s.id_user in participants else None

        result_list.append(
            {
                "id": s.id_tasks_to_participant,
                "task_id": s.id_tasks,
                "user_id": s.id_user,
                "task_title": task.name_tasks if task else "Удаленное задание",
                "user_name": user.name_user if user else "Неизвестный участник",
                "user_institute": (
                    participant.institute_participant if participant else "Не указан"
                ),
                "submission_data": s.url_participant,
                "status": s.status,
                "score": s.check.points_check_tasks if s.check else 0,
                "max_points": task.points_per_action if task else 0,
                "submitted_at": (
                    s.date_task_to_participant.isoformat()
                    if s.date_task_to_participant
                    else None
                ),
            }
        )

    return result_list


@app.post("/api/submissions/{sub_id}/review", tags=["Admin/Reviews"])
async def review_submission(
    sub_id: int,
    status: str = Form(...),
    score: float = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Выставление оценки работе"""
    result = await db.execute(
        select(TasksToParticipant)
        .options(selectinload(TasksToParticipant.check))
        .where(TasksToParticipant.id_tasks_to_participant == sub_id)
    )
    submission = result.scalar_one_or_none()

    if not submission:
        raise HTTPException(status_code=404, detail="Работа не найдена.")

    # Защита от повторной проверки другим администратором
    if (
        submission.check
        and submission.check.reviewed_by_id
        and submission.check.reviewed_by_id != current_user.id_user
        and submission.status in ("approved", "rejected")
    ):
        raise HTTPException(
            status_code=403,
            detail="Эта работа уже проверена другим администратором и не может быть изменена.",
        )

    # Получаем задание для проверки макс. балла
    task_result = await db.execute(
        select(Task).where(Task.id_tasks == submission.id_tasks)
    )
    task = task_result.scalar_one_or_none()

    if task and score > task.points_per_action:
        raise HTTPException(
            status_code=400,
            detail=f"Оценка не может превышать максимальный балл задания ({task.points_per_action})",
        )

    submission.status = status

    # Обновляем или создаём запись проверки
    if submission.check:
        submission.check.points_check_tasks = int(score)
        submission.check.reviewed_by_id = current_user.id_user
    else:
        check = CheckTasks(
            id_submission=sub_id,
            id_user=current_user.id_user,
            id_tasks=submission.id_tasks,
            points_check_tasks=int(score),
            reviewed_by_id=current_user.id_user,
        )
        db.add(check)

    # Уведомление пользователю
    status_text = (
        "принята"
        if status == "approved"
        else "отклонена" if status == "rejected" else "проверена"
    )
    notification = Notification(
        id_user=submission.id_user,
        title="📝 Работа проверена",
        message=f"Ваша работа по заданию #{submission.id_tasks} {status_text}. Оценка: {score} баллов.",
    )
    db.add(notification)

    await db.commit()
    return {"message": "Оценка успешно сохранена"}


# ========== УВЕДОМЛЕНИЯ ==========
@app.get("/api/users/me/notifications", tags=["User/Notifs"])
async def get_user_notifications(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_user),
):
    """Получение уведомлений пользователя"""
    result = await db.execute(
        select(Notification)
        .where(Notification.id_user == current_user.id_user)
        .order_by(Notification.id_notification.desc())
        .limit(50)
    )
    notifs = result.scalars().all()

    return [
        {
            "id": n.id_notification,
            "title": n.title,
            "message": n.message,
            "is_read": n.is_read,
            "created_at": n.created_at,
        }
        for n in notifs
    ]


@app.get("/api/users/me/notifications/unread-count", tags=["User/Notifs"])
async def get_unread_count(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_user),
):
    """Количество непрочитанных уведомлений"""
    result = await db.execute(
        select(func.count(Notification.id_notification)).where(
            Notification.id_user == current_user.id_user, Notification.is_read == False
        )
    )
    count = result.scalar() or 0
    return {"unread_count": count}


@app.post("/api/notifications/{n_id}/read", tags=["User/Notifs"])
async def mark_notification_read(
    n_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_user),
):
    """Отметить уведомление как прочитанное"""
    await db.execute(
        update(Notification)
        .where(
            Notification.id_notification == n_id,
            Notification.id_user == current_user.id_user,
        )
        .values(is_read=True)
    )
    await db.commit()
    return {"status": "ok"}


@app.post("/api/notifications/mark-all-read", tags=["User/Notifs"])
async def mark_all_notifications_read(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_user),
):
    """Отметить все уведомления как прочитанные"""
    await db.execute(
        update(Notification)
        .where(
            Notification.id_user == current_user.id_user, Notification.is_read == False
        )
        .values(is_read=True)
    )
    await db.commit()
    return {"status": "ok"}


# ========== СТАТИСТИКА И ОТЧЁТЫ ==========
@app.get("/api/stats/dashboard", tags=["Admin/Stats"])
async def get_dashboard_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Статистика для дашборда"""
    result = await db.execute(
        select(func.count(User.id_user)).where(User.role_user == "participant")
    )
    total_users = result.scalar() or 0

    result = await db.execute(
        select(
            Participant.institute_participant, func.count(Participant.id_participant)
        )
        .join(User, User.id_user == Participant.id_user)
        .where(User.role_user == "participant")
        .group_by(Participant.institute_participant)
    )
    inst_stats = result.all()

    # Статистика по заданиям: сколько участников выполнили каждое
    result = await db.execute(
        select(
            Task.id_tasks,
            Task.name_tasks,
            func.count(TasksToParticipant.id_tasks_to_participant).label("completed"),
        )
        .outerjoin(TasksToParticipant, TasksToParticipant.id_tasks == Task.id_tasks)
        .group_by(Task.id_tasks, Task.name_tasks)
        .order_by(Task.id_tasks.desc())
    )
    tasks_stats = result.all()

    return {
        "total_users": total_users,
        "institutes_activity": [{"institute": i, "count": c} for i, c in inst_stats],
        "tasks_stats": [
            {"task_id": tid, "title": name, "completed_count": cnt}
            for tid, name, cnt in tasks_stats
        ],
    }


@app.get("/api/reports/contacts/txt", tags=["Admin/Stats"])
async def export_contacts_txt(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Выгрузка контактных данных участников в TXT"""
    result = await db.execute(select(User).where(User.role_user == "participant"))
    users = result.scalars().all()

    lines = [
        "КОНТАКТЫ УЧАСТНИКОВ МЕДИАКОНКУРСА\n",
        "=" * 50 + "\n",
        f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n",
        f"Всего участников: {len(users)}\n",
        "=" * 50 + "\n\n",
    ]

    for i, user in enumerate(users, 1):
        part_result = await db.execute(
            select(Participant).where(Participant.id_user == user.id_user)
        )
        participant = part_result.scalar_one_or_none()
        lines.append(f"{i}. {user.name_user}\n")
        lines.append(f"   Институт: {participant.institute_participant if participant else '—'}\n")
        lines.append(f"   Email:    {user.email_user}\n")
        vk = participant.vk_participant if participant and participant.vk_participant else "—"
        lines.append(f"   ВКонтакте: {vk}\n")
        tg = participant.tg_participant if participant and participant.tg_participant else "—"
        lines.append(f"   Telegram:  {tg}\n")
        lines.append("\n")

    return PlainTextResponse(
        "".join(lines),
        headers={"Content-Disposition": "attachment; filename=contacts_media.txt"},
    )


@app.get("/api/reports/contacts/xlsx", tags=["Admin/Stats"])
async def export_contacts_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Выгрузка контактных данных участников в XLSX"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    result = await db.execute(select(User).where(User.role_user == "participant"))
    users = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Контакты участников"

    header_fill = PatternFill(start_color="233F26", end_color="233F26", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="DDDDDD"),
    )

    headers = ["№", "ФИО", "Институт", "Email", "ВКонтакте", "Telegram"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 28

    for row_idx, user in enumerate(users, 2):
        part_result = await db.execute(
            select(Participant).where(Participant.id_user == user.id_user)
        )
        participant = part_result.scalar_one_or_none()

        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=user.name_user)
        ws.cell(row=row_idx, column=3, value=participant.institute_participant if participant else "—")
        ws.cell(row=row_idx, column=4, value=user.email_user)
        ws.cell(row=row_idx, column=5, value=participant.vk_participant if participant and participant.vk_participant else "—")
        ws.cell(row=row_idx, column=6, value=participant.tg_participant if participant and participant.tg_participant else "—")

        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = thin_border

    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contacts_media.xlsx"},
    )


@app.get("/api/reports/csv", tags=["Admin/Stats"])
async def export_csv(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Экспорт рейтинга в CSV"""
    result = await db.execute(select(User).where(User.role_user == "participant"))
    users = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["ФИО", "Институт", "Суммарный балл"])

    for user in users:
        part_result = await db.execute(
            select(Participant).where(Participant.id_user == user.id_user)
        )
        participant = part_result.scalar_one_or_none()

        score_result = await db.execute(
            select(func.sum(CheckTasks.points_check_tasks))
            .select_from(TasksToParticipant)
            .join(
                CheckTasks,
                TasksToParticipant.id_tasks_to_participant == CheckTasks.id_submission,
            )
            .where(TasksToParticipant.id_user == user.id_user)
        )
        total_score = score_result.scalar() or 0

        writer.writerow(
            [
                user.name_user,
                participant.institute_participant if participant else "Не указан",
                total_score,
            ]
        )

    return Response(
        content=output.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=rating_media.csv"},
    )


@app.get("/api/reports/txt", tags=["Admin/Stats"])
async def export_txt(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Экспорт рейтинга в TXT"""
    result = await db.execute(select(User).where(User.role_user == "participant"))
    users = result.scalars().all()

    user_scores = []
    for user in users:
        part_result = await db.execute(
            select(Participant).where(Participant.id_user == user.id_user)
        )
        participant = part_result.scalar_one_or_none()

        score_result = await db.execute(
            select(func.sum(CheckTasks.points_check_tasks))
            .select_from(TasksToParticipant)
            .join(
                CheckTasks,
                TasksToParticipant.id_tasks_to_participant == CheckTasks.id_submission,
            )
            .where(TasksToParticipant.id_user == user.id_user)
        )
        total_score = score_result.scalar() or 0

        user_scores.append(
            (
                user.name_user,
                participant.institute_participant if participant else "?",
                total_score,
            )
        )

    user_scores.sort(key=lambda x: x[2], reverse=True)

    lines = [
        "РЕЙТИНГ УЧАСТНИКОВ МЕДИАКОНКУРСА\n",
        "=" * 50 + "\n",
        f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n",
        f"Всего участников: {len(users)}\n",
        f"Общая сумма баллов: {sum(s[2] for s in user_scores)}\n",
        "=" * 50 + "\n\n",
        "ТОП УЧАСТНИКОВ:\n",
        "-" * 40 + "\n",
    ]

    for i, (name, inst, score) in enumerate(user_scores[:20], 1):
        lines.append(f"{i}. {name} ({inst}) — {score} баллов\n")

    return PlainTextResponse(
        "".join(lines),
        headers={"Content-Disposition": "attachment; filename=rating_media.txt"},
    )


@app.get("/api/reports/xlsx", tags=["Admin/Stats"])
async def export_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Экспорт рейтинга в XLSX с автоподбором ширины столбцов"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    result = await db.execute(select(User).where(User.role_user == "participant"))
    users = result.scalars().all()

    user_scores = []
    for user in users:
        part_result = await db.execute(
            select(Participant).where(Participant.id_user == user.id_user)
        )
        participant = part_result.scalar_one_or_none()

        score_result = await db.execute(
            select(func.sum(CheckTasks.points_check_tasks))
            .select_from(TasksToParticipant)
            .join(
                CheckTasks,
                TasksToParticipant.id_tasks_to_participant == CheckTasks.id_submission,
            )
            .where(TasksToParticipant.id_user == user.id_user)
        )
        total_score = score_result.scalar() or 0
        user_scores.append((
            user.name_user,
            participant.institute_participant if participant else "Не указан",
            total_score,
        ))

    user_scores.sort(key=lambda x: x[2], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Рейтинг участников"

    # Стили заголовков
    header_fill = PatternFill(start_color="233F26", end_color="233F26", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="DDDDDD"),
    )

    headers = ["№", "ФИО", "Институт", "Баллы"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 28

    # Данные
    for row_idx, (name, inst, score) in enumerate(user_scores, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=inst)
        score_cell = ws.cell(row=row_idx, column=4, value=score)
        score_cell.alignment = Alignment(horizontal="center")
        score_cell.font = Font(bold=True)
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Автоподбор ширины по содержимому
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Закрепить шапку
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rating_media.xlsx"},
    )

    # ========== РЕЙТИНГ ==========


@app.get("/api/rating", tags=["User"])
async def get_rating(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_user),
):
    """Получение рейтинга пользователя"""

    # Получаем всех участников с их баллами
    result = await db.execute(
        select(
            User,
            func.coalesce(func.sum(CheckTasks.points_check_tasks), 0).label(
                "total_score"
            ),
        )
        .outerjoin(TasksToParticipant, TasksToParticipant.id_user == User.id_user)
        .outerjoin(
            CheckTasks,
            CheckTasks.id_submission == TasksToParticipant.id_tasks_to_participant,
        )
        .where(User.role_user == "participant")
        .group_by(User.id_user)
        .order_by(func.coalesce(func.sum(CheckTasks.points_check_tasks), 0).desc())
    )

    ratings = []
    for user, score in result.all():
        # Получаем институт участника
        part_result = await db.execute(
            select(Participant).where(Participant.id_user == user.id_user)
        )
        participant = part_result.scalar_one_or_none()

        ratings.append(
            {
                "user_id": user.id_user,
                "full_name": user.name_user,
                "institute": (
                    participant.institute_participant if participant else "Не указан"
                ),
                "total_score": int(score),
            }
        )

    # Находим место текущего пользователя
    user_rank = None
    for i, r in enumerate(ratings, 1):
        if r["user_id"] == current_user.id_user:
            user_rank = i
            break

    return {
        "ratings": ratings,
        "user_rank": user_rank,
        "user_score": next(
            (r["total_score"] for r in ratings if r["user_id"] == current_user.id_user),
            0,
        ),
    }


# ========== О КОНКУРСЕ ==========

@app.get("/api/about", tags=["About"])
async def get_about(db: AsyncSession = Depends(get_db)):
    """Информация о конкурсе + новости (публичный)"""
    result = await db.execute(select(AboutContest).where(AboutContest.id == 1))
    about = result.scalar_one_or_none()

    result = await db.execute(select(News).order_by(News.created_at.desc()))
    news_list = result.scalars().all()

    return {
        "content": about.content if about else "",
        "news": [
            {
                "id": n.id,
                "title": n.title,
                "content": n.content,
                "created_at": n.created_at.isoformat() if n.created_at else None,
            }
            for n in news_list
        ],
    }


@app.put("/api/admin/about", tags=["About"])
async def update_about(
    content: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Обновить текст о конкурсе"""
    result = await db.execute(select(AboutContest).where(AboutContest.id == 1))
    about = result.scalar_one_or_none()

    if about:
        about.content = content
        about.updated_at = datetime.utcnow()
        about.updated_by = current_user.id_user
    else:
        db.add(AboutContest(id=1, content=content, updated_by=current_user.id_user))

    await db.commit()
    return {"message": "Обновлено"}


@app.post("/api/admin/news", tags=["About"])
async def create_news(
    title: str = Form(...),
    content: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Создать новость (рассылает уведомления всем участникам)"""
    news = News(title=title, content=content, created_by=current_user.id_user)
    db.add(news)
    await db.flush()

    # Уведомления всем участникам
    result = await db.execute(select(User).where(User.role_user == "participant"))
    for user in result.scalars().all():
        db.add(Notification(
            id_user=user.id_user,
            title=f"Новость: {title}",
            message=content[:200] + ("..." if len(content) > 200 else ""),
        ))

    await db.commit()
    return {"message": "Новость опубликована", "id": news.id}


@app.put("/api/admin/news/{news_id}", tags=["About"])
async def update_news(
    news_id: int,
    title: str = Form(...),
    content: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Редактировать новость"""
    result = await db.execute(select(News).where(News.id == news_id))
    news = result.scalar_one_or_none()
    if not news:
        raise HTTPException(status_code=404, detail="Новость не найдена.")
    news.title = title
    news.content = content
    news.updated_at = datetime.utcnow()
    await db.commit()
    return {"message": "Обновлено"}


@app.delete("/api/admin/news/{news_id}", tags=["About"])
async def delete_news(
    news_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    """Удалить новость"""
    result = await db.execute(select(News).where(News.id == news_id))
    news = result.scalar_one_or_none()
    if not news:
        raise HTTPException(status_code=404, detail="Новость не найдена.")
    await db.delete(news)
    await db.commit()
    return {"message": "Удалено"}


# ========== FAQ ==========

@app.get("/api/faq", tags=["FAQ"])
async def get_faq(db: AsyncSession = Depends(get_db)):
    """Список вопросов FAQ (публичный)"""
    result = await db.execute(select(FAQ).order_by(FAQ.order_num, FAQ.id))
    items = result.scalars().all()
    return [{"id": f.id, "question": f.question, "answer": f.answer} for f in items]


@app.post("/api/admin/faq", tags=["FAQ"])
async def create_faq(
    question: str = Form(...),
    answer: str = Form(...),
    order_num: int = Form(0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    item = FAQ(question=question, answer=answer, order_num=order_num, created_by=current_user.id_user)
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return {"id": item.id, "message": "Вопрос добавлен"}


@app.put("/api/admin/faq/{faq_id}", tags=["FAQ"])
async def update_faq(
    faq_id: int,
    question: str = Form(...),
    answer: str = Form(...),
    order_num: int = Form(0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    result = await db.execute(select(FAQ).where(FAQ.id == faq_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Вопрос не найден.")
    item.question = question
    item.answer = answer
    item.order_num = order_num
    await db.commit()
    return {"message": "Обновлено"}


@app.delete("/api/admin/faq/{faq_id}", tags=["FAQ"])
async def delete_faq(
    faq_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(auth.get_current_admin),
):
    result = await db.execute(select(FAQ).where(FAQ.id == faq_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Вопрос не найден.")
    await db.delete(item)
    await db.commit()
    return {"message": "Удалено"}


# ========== СТАТИКА ==========
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
app.mount("/", StaticFiles(directory=".", html=True), name="static")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8001)
